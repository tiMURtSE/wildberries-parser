from typing import List
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from basic_decor_library.Workbook import Workbook

from models.Product.Product import Product
from models.Review.Review import Review

class ResultWorkbook(Workbook):
    RESULT_WORKBOOK_FILE_PATH = "C:/Users/user10/Desktop/Stuff/Плюсы и минусы товара/Wildberries/Отзывы из Wildberries.xlsx"

    PRODUCT_ID_COL_TITLE = "product_id"
    ARTICLE_COL_TITLE = "c:vendor_sku"
    CUSTOMER_NAME_COL_TITLE = "customer_name"
    COMMENT_COL_TITLE = "review"
    RATE_COL_TITLE = "rate"

    def __init__(self):
        self._workbook = openpyxl.load_workbook(self.RESULT_WORKBOOK_FILE_PATH)
        self._sheet: Worksheet = self._workbook.active

    def write_result(self, product: Product):
        row_start_position = self._get_row_start_position()

        for index, review in enumerate(product.reviews):
            row_index = index + row_start_position

            self._write_review(row_index=row_index, review=review, product=product)

        self._workbook.save(self.RESULT_WORKBOOK_FILE_PATH)

    def _write_review(self, row_index: int, review: Review, product: Product):
        product_id_col_index = self.find_column_index(column_name=self.PRODUCT_ID_COL_TITLE)
        article_col_index = self.find_column_index(column_name=self.ARTICLE_COL_TITLE)
        customer_name_col_index = self.find_column_index(column_name=self.CUSTOMER_NAME_COL_TITLE)
        comment_col_index = self.find_column_index(column_name=self.COMMENT_COL_TITLE)
        rate_col_index = self.find_column_index(column_name=self.RATE_COL_TITLE)

        self._sheet.cell(row=row_index, column=product_id_col_index).value = product.id
        self._sheet.cell(row=row_index, column=article_col_index).value = product.article

        self._sheet.cell(row=row_index, column=customer_name_col_index).value = review._customer_name
        self._sheet.cell(row=row_index, column=comment_col_index).value = review._comment
        self._sheet.cell(row=row_index, column=rate_col_index).value = review._rate

    def _get_row_start_position(self):
        return self._sheet.max_row + 1